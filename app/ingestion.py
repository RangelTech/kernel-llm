"""File ingestion: extract text, chunk, embed, store vectors.

Triggered by POST /v1/ingest-file (Cloud Tasks in production, an inline
background task in dev). Status transitions are written to the files table
so the frontend can poll progress."""

import io
import logging

from app.embeddings import embed_texts
from app.storage import load_payload
from app.trace import _get_pool

logger = logging.getLogger(__name__)

CHUNK_CHARS = 1200
CHUNK_OVERLAP = 150


def extract_text(data: bytes, content_type: str, name: str) -> str:
    lower = name.lower()
    if lower.endswith(".pdf") or content_type == "application/pdf":
        import fitz  # pymupdf

        with fitz.open(stream=data, filetype="pdf") as doc:
            return "\n\n".join(page.get_text() for page in doc)
    if lower.endswith(".docx"):
        from docx import Document

        document = Document(io.BytesIO(data))
        return "\n".join(p.text for p in document.paragraphs)
    if lower.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        parts = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    parts.append(" | ".join(cells))
        return "\n".join(parts)
    # txt, md, csv and anything text-like.
    return data.decode("utf-8", errors="replace")


def chunk_text(text: str) -> list[str]:
    text = " ".join(text.split())  # collapse whitespace
    if not text:
        return []
    chunks = []
    start = 0
    while start < len(text):
        end = start + CHUNK_CHARS
        chunks.append(text[start:end])
        start = end - CHUNK_OVERLAP
    return chunks


async def _set_status(file_id: str, status: str, *, error: str = "", chunks: int | None = None):
    pool = await _get_pool()
    async with pool.connection() as conn:
        await conn.execute(
            """UPDATE files SET status = %s, error_detail = %s,
                      chunk_count = COALESCE(%s, chunk_count), updated_at = now()
                WHERE id = %s""",
            (status, error or None, chunks, file_id),
        )


async def ingest_file(file_id: str, embedding_spec: dict) -> None:
    pool = await _get_pool()
    async with pool.connection() as conn:
        cursor = await conn.execute("SELECT * FROM files WHERE id = %s", (file_id,))
        row = await cursor.fetchone()
        if row is None:
            logger.warning("ingest: file %s not found", file_id)
            return
        columns = [d.name for d in cursor.description]
        file = dict(zip(columns, row, strict=True))

    await _set_status(file_id, "processing")
    try:
        data = load_payload(file["storage_path"])
        text = extract_text(data, file["content_type"], file["name"])
        chunks = chunk_text(text)
        if not chunks:
            raise ValueError("arquivo sem texto extraível")

        # Embed in batches to respect API limits.
        vectors: list[list[float]] = []
        for i in range(0, len(chunks), 32):
            vectors.extend(await embed_texts(embedding_spec, chunks[i : i + 32]))

        async with pool.connection() as conn:
            await conn.execute("DELETE FROM file_chunks WHERE file_id = %s", (file_id,))
            for index, (content, vector) in enumerate(zip(chunks, vectors, strict=True)):
                await conn.execute(
                    """INSERT INTO file_chunks
                           (file_id, tenant_id, chunk_index, content, embedding)
                       VALUES (%s, %s, %s, %s, %s)""",
                    (file_id, file["tenant_id"], index, content, str(vector)),
                )
        await _set_status(file_id, "ready", chunks=len(chunks))
        logger.info("ingested file %s: %d chunks", file_id, len(chunks))
    except Exception as exc:  # noqa: BLE001 — status must reflect the failure
        logger.exception("ingestion failed for %s", file_id)
        await _set_status(file_id, "error", error=str(exc)[:500])


async def search_chunks(
    embedding_spec: dict, question: str, file_ids: list[str], tenant_id, top_k: int = 5
) -> list[dict]:
    """Vector search restricted to the given files (the agent's RAG scope)."""
    if not file_ids:
        return []
    [vector] = await embed_texts(embedding_spec, [question])
    pool = await _get_pool()
    async with pool.connection() as conn:
        cursor = await conn.execute(
            """SELECT c.content, c.chunk_index, f.name,
                      1 - (c.embedding <=> %s::vector) AS similarity
                 FROM file_chunks c
                 JOIN files f ON f.id = c.file_id
                WHERE c.file_id = ANY(%s::uuid[]) AND c.tenant_id = %s
                ORDER BY c.embedding <=> %s::vector
                LIMIT %s""",
            (str(vector), file_ids, tenant_id, str(vector), top_k),
        )
        rows = await cursor.fetchall()
    return [
        {"file": r[2], "chunk": r[1], "similarity": round(float(r[3]), 3), "content": r[0]}
        for r in rows
    ]
